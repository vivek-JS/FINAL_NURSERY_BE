#!/usr/bin/env python3
import openpyxl
import json
import sys
from datetime import datetime

file_path = "middlewares/soft_booking_2.xlsx"

column_mapping = {
    0:  'Date',
    1:  'Booking NO.',
    2:  'Name',
    3:  'Mobile No.',
    4:  'Address',
    5:  'Taluka',
    6:  'District',
    7:  'Advance On Booking Receipts',
    8:  'adv match or not',
    9:  'Advance Amt.',
    10: 'Crop',
    11: 'Variety',
    12: 'Media',
    13: 'Expected Nursery',
    14: 'Plant Qty.',
    15: 'Rate',
    16: 'Expected Del. Date',
    17: 'Old Del. Date',
    18: 'Del. Y/N',
    19: 'Actually Del. Date',
    20: 'Invoice amount',
    21: 'Bal. Amt.',
    22: 'Refrence',
    23: 'Order By',
    24: 'Ad. Amt. Mode',
    25: 'Bank',
    26: 'CH No.',
    27: 'Advance Date',
    28: 'Receipt Code',
    29: 'ADV Y/N',
    30: 'CC Y/N',
    31: 'Remark',
}

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['Sheet2']

    rows = []
    for i in range(2, ws.max_row + 1):
        row_data = {}
        for j, cell in enumerate(ws[i]):
            value = cell.value
            if value is not None:
                if isinstance(value, datetime):
                    value = value.strftime("%m/%d/%y")
                elif isinstance(value, bool):
                    value = value
                elif isinstance(value, (int, float)):
                    value = value
                else:
                    value = str(value)

                key = column_mapping.get(j)
                if key:
                    row_data[key] = value

        # Skip blank rows (no meaningful values at all)
        if not any(v for v in row_data.values() if v not in (None, '', False, 0)):
            break

        # Normalize known variety name mismatches
        if row_data.get('Variety') == 'G-9':
            row_data['Variety'] = 'G9'

        rows.append(row_data)

    print(json.dumps(rows, indent=2))

except Exception as e:
    print(f"Error reading file: {e}", file=sys.stderr)
    sys.exit(1)
