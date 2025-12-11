#!/usr/bin/env python3
"""
Read password-protected Excel file from deployment folder
Password: AV1312
"""

import openpyxl
import msoffcrypto
import json
import sys
import os
from datetime import datetime
from pathlib import Path
import io

password = "AV1312"
excel_file = "deployment/BOOKING DETAILS 2025-26 (7).xlsx"

def read_excel_file():
    try:
        # Get absolute path
        script_dir = Path(__file__).parent
        file_path = script_dir / excel_file
        
        if not file_path.exists():
            print(f"❌ Excel file not found at: {file_path}")
            return
        
        print(f"📖 Reading Excel file: {file_path}")
        print("━" * 50)
        print()
        
        # Step 1: Decrypt the password-protected file
        print("🔓 Decrypting password-protected file...")
        with open(file_path, "rb") as file:
            encrypted_file = msoffcrypto.OfficeFile(file)
            encrypted_file.load_key(password=password)
            
            decrypted_file = io.BytesIO()
            encrypted_file.decrypt(decrypted_file)
            print("✅ File decrypted successfully!")
            print()
            
            # Step 2: Load workbook
            decrypted_file.seek(0)  # Reset to beginning
            workbook = openpyxl.load_workbook(decrypted_file, data_only=True)
            
            # Display workbook information
            print(f"📊 Workbook Information:")
            print(f"─────────────────────────────────────────")
            print(f"Total Sheets: {len(workbook.sheetnames)}")
            print(f"Sheet Names: {', '.join(workbook.sheetnames)}")
            print()
            
            # Process each sheet
            for sheet_index, sheet_name in enumerate(workbook.sheetnames):
                print(f"\n📄 Sheet {sheet_index + 1}: \"{sheet_name}\"")
                print("═" * 50)
                
                sheet = workbook[sheet_name]
                
                print(f"Total Rows: {sheet.max_row}")
                print(f"Total Columns: {sheet.max_column}")
                print()
                
                if sheet.max_row <= 1:
                    print("⚠️  Sheet appears to be empty\n")
                    continue
                
                # Get headers from first row
                headers = []
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=1, column=col)
                    header = cell.value if cell.value else f"Column_{col}"
                    headers.append(str(header))
                
                print(f"📋 Column Headers ({len(headers)}):")
                for idx, header in enumerate(headers, 1):
                    print(f"   {idx}. {header}")
                print()
                
                # Find Name column index
                name_header = None
                for header in headers:
                    if header.lower().strip() == "name":
                        name_header = header
                        break
                
                if not name_header:
                    print("⚠️  No 'Name' column found in this sheet\n")
                    continue
                
                print(f"✅ Found 'Name' column: '{name_header}'")
                print(f"📊 Processing all rows (filtering rows with Name)...")
                
                # Convert rows to dictionaries - only include rows where Name is present
                rows = []
                original_row_numbers = []  # Track original Excel row numbers
                name_col_idx = headers.index(name_header) + 1
                
                for row_idx in range(2, sheet.max_row + 1):  # Process all rows
                    # Check if Name column has a value
                    name_cell = sheet.cell(row=row_idx, column=name_col_idx)
                    name_value = name_cell.value
                    
                    # Skip rows where Name is empty, None, or just whitespace
                    if name_value is None or (isinstance(name_value, str) and not name_value.strip()):
                        continue
                    
                    row_data = {}
                    
                    for col_idx, header in enumerate(headers, 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        value = cell.value
                        
                        if value is not None:
                            # Convert datetime to string
                            if isinstance(value, datetime):
                                value = value.strftime("%Y-%m-%d")
                            elif isinstance(value, (int, float)):
                                value = value
                            else:
                                value = str(value).strip()
                        
                        row_data[header] = value if value is not None and value != "" else None
                    
                    # Add original Excel row number to the data
                    row_data['_ExcelRowNumber'] = row_idx
                    row_data['_FilteredRowNumber'] = len(rows) + 1
                    
                    rows.append(row_data)
                    original_row_numbers.append(row_idx)  # Store original row number
                
                # Summary statistics
                print(f"\n📈 Summary:")
                print(f"   Total Rows Processed (all rows with Name): {len(rows)}")
                print(f"   Total Rows in Sheet: {sheet.max_row - 1}")
                skipped_rows = (sheet.max_row - 1) - len(rows)
                if skipped_rows > 0:
                    print(f"   Rows Skipped (no Name): {skipped_rows}")
                
                if len(rows) == 0:
                    print("\n⚠️  No rows found with Name field\n")
                    continue
                
                # Display first and last row information
                if len(rows) > 0 and len(original_row_numbers) > 0:
                    first_excel_row = original_row_numbers[0]
                    last_excel_row = original_row_numbers[-1]
                    first_row_data = rows[0]
                    last_row_data = rows[-1]
                    
                    print(f"\n🔍 First and Last Row Detection:")
                    print("═" * 50)
                    print(f"\n📌 FIRST ROW (Filtered Row #1 / Excel Row #{first_excel_row}):")
                    print(f"   Name: {first_row_data.get(name_header, 'N/A')}")
                    for header in ['Date', 'Booking NO.', 'Mobile No.', 'Address', 'Taluka', 'District']:
                        if header in first_row_data:
                            value = first_row_data.get(header)
                            if value:
                                display_value = str(value)
                                if len(display_value) > 60:
                                    display_value = display_value[:60] + "..."
                                print(f"   {header}: {display_value}")
                    
                    print(f"\n📌 LAST ROW (Filtered Row #{len(rows)} / Excel Row #{last_excel_row}):")
                    print(f"   Name: {last_row_data.get(name_header, 'N/A')}")
                    for header in ['Date', 'Booking NO.', 'Mobile No.', 'Address', 'Taluka', 'District']:
                        if header in last_row_data:
                            value = last_row_data.get(header)
                            if value:
                                display_value = str(value)
                                if len(display_value) > 60:
                                    display_value = display_value[:60] + "..."
                                print(f"   {header}: {display_value}")
                    
                    print(f"\n📊 Row Range:")
                    print(f"   First Excel Row: {first_excel_row}")
                    print(f"   Last Excel Row: {last_excel_row}")
                    print(f"   Total Excel Rows Span: {last_excel_row - first_excel_row + 1}")
                    print(f"   Filtered Rows Count: {len(rows)}")
                    print(f"   Rows between first and last: {last_excel_row - first_excel_row - 1}")
                
                # Display first 5 rows as sample
                print(f"\n📝 First 5 Rows (Sample):")
                print("─────────────────────────────────────────")
                for row_idx, row in enumerate(rows[:5], 1):
                    print(f"\nRow {row_idx}:")
                    for header in headers[:10]:  # Show first 10 columns
                        value = row.get(header)
                        if value is not None:
                            display_value = str(value)
                            if len(display_value) > 50:
                                display_value = display_value[:50] + "..."
                        else:
                            display_value = "(empty)"
                        print(f"   {header}: {display_value}")
                    if len(headers) > 10:
                        print(f"   ... ({len(headers) - 10} more columns)")
                
                # Count non-empty values per column
                print(f"\n   Non-empty values per column (first 15):")
                for header in headers[:15]:
                    non_empty = sum(1 for row in rows if row.get(header) not in [None, ""])
                    if len(rows) > 0:
                        percentage = (non_empty / len(rows)) * 100
                        print(f"   {header}: {non_empty}/{len(rows)} ({percentage:.1f}%)")
                
                # Save to JSON file
                json_output = script_dir / f"deployment/excel-data-{sheet_name.replace('/', '_')}-with-name-only.json"
                with open(json_output, 'w', encoding='utf-8') as f:
                    json.dump(rows, f, indent=2, ensure_ascii=False, default=str)
                print(f"\n💾 Filtered data (rows with Name) saved to: {json_output}")
                print(f"   Total rows in JSON: {len(rows)}")
            
            print("\n" + "━" * 50)
            print("✅ Excel file reading completed!")
            
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    read_excel_file()

