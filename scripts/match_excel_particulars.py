#!/usr/bin/env python3
"""
Match Excel 'Particulars' rows to prod farmers + orders (fuzzy).
Loads PROD_MONGO_URL from ../.env — do not print the URI.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

from dotenv import dotenv_values
from openpyxl import load_workbook
from pymongo import MongoClient
from rapidfuzz import fuzz, process

ROOT = Path(__file__).resolve().parents[1]
ENV_PATH = ROOT / ".env"
XLSX_DEFAULT = Path.home() / "Downloads" / "New Microsoft Excel Worksheet (2).xlsx"

# Thresholds (0–100), per user guidance
MIN_FARMER = 85
MIN_VILLAGE = 60
MIN_PLANT = 50


def norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def tsv_cell(s) -> str:
    if s is None:
        return ""
    t = str(s).replace("\t", " ").replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", t).strip()


def amounts_from_row(row: tuple | None) -> dict[str, object]:
    """Debit / Credit from source sheet (cols B & C), same as original form."""
    if not row:
        return {"excel_debit": "", "excel_credit": ""}
    debit = row[1] if len(row) > 1 else None
    credit = row[2] if len(row) > 2 else None
    return {
        "excel_debit": "" if debit is None else debit,
        "excel_credit": "" if credit is None else credit,
    }


def load_uri() -> str:
    env = dotenv_values(ENV_PATH)
    uri = env.get("PROD_MONGO_URL") or env.get("MONGO_URL")
    if not uri:
        print("Missing PROD_MONGO_URL (or MONGO_URL) in .env", file=sys.stderr)
        sys.exit(1)
    return uri.strip()


def parse_tokens(line: str) -> list[str]:
    line = (line or "").strip()
    if not line:
        return []
    tokens = [t for t in re.split(r"\s+", line) if t]
    while tokens and re.sub(r"\.+$", "", tokens[-1].lower()) in ("adv", "advv"):
        tokens.pop()
    return tokens


def iter_splits(tokens: list[str]):
    """Yield (farmer_tokens, village_tokens, plant_tokens). plant may be empty."""
    n = len(tokens)
    if n == 0:
        return
    for plant_k in range(0, min(5, n) + 1):
        rest = n - plant_k
        if rest < 1:
            continue
        plant_toks = tokens[n - plant_k :] if plant_k else []
        mid = tokens[: n - plant_k]
        for village_m in range(1, min(4, len(mid)) + 1):
            farmer_toks = mid[:-village_m]
            village_toks = mid[-village_m:]
            if not farmer_toks:
                continue
            yield (farmer_toks, village_toks, plant_toks)


def best_plant_score(plant_phrase: str, plant_strings: list[str]) -> float:
    if not plant_phrase.strip():
        return 0.0
    p = norm(plant_phrase)
    best = 0.0
    for s in plant_strings:
        if not s:
            continue
        sn = norm(s)
        best = max(
            best,
            float(fuzz.ratio(p, sn)),
            float(fuzz.partial_ratio(p, sn)),
            float(fuzz.token_sort_ratio(p, sn)),
        )
    return best


OUTPUT_KEYS = [
    "excel_line",
    "excel_debit",
    "excel_credit",
    "status",
    "farmer_parse",
    "village_parse",
    "plant_parse",
    "score_farmer",
    "score_village_parse_vs_db_village",
    "score_village_parse_vs_farmer_village",
    "score_plant",
    "matched_farmer_name",
    "matched_village_db",
    "orderId",
    "order_plant",
    "order_subtype",
    "numberOfPlants",
    "orderStatus",
    "weighted_total",
    "note",
]


def run_match(xlsx: Path) -> list[dict]:
    if not xlsx.is_file():
        raise FileNotFoundError(f"Excel not found: {xlsx}")

    uri = load_uri()
    client = MongoClient(uri, serverSelectionTimeoutMS=20000)
    db = client.get_database("nursery")

    # Prod often stores village as plain string on Farmer; Village collection may be empty.
    villages = {str(v["_id"]): v.get("village") or "" for v in db.villages.find({}, {"village": 1})}
    farmers = list(db.farmers.find({}, {"name": 1, "village": 1}))
    farmer_by_id: dict[str, dict] = {}
    village_from_farmers: set[str] = set()
    for f in farmers:
        vraw = f.get("village")
        if isinstance(vraw, str) and vraw.strip():
            village_from_farmers.add(vraw.strip())
        elif vraw is not None:
            village_from_farmers.add(str(vraw).strip())
        vid = str(vraw or "")
        vname = villages.get(vid, "") if vid in villages else (vraw if isinstance(vraw, str) else "")
        farmer_by_id[str(f["_id"])] = {
            "_id": f["_id"],
            "name": (f.get("name") or "").strip(),
            "village_name": (vname or (vraw if isinstance(vraw, str) else "") or "").strip(),
        }
    village_names_list = sorted(village_from_farmers | {v for v in villages.values() if v})
    village_names_norm = [norm(v) for v in village_names_list]

    plant_strings: list[str] = []
    subtype_by_id: dict[str, tuple[str, str]] = {}
    for doc in db.plantcms.find({}, {"name": 1, "subtypes": 1}):
        pname = doc.get("name") or ""
        if pname:
            plant_strings.append(pname)
        for st in doc.get("subtypes") or []:
            stid = str(st.get("_id"))
            stname = st.get("name") or ""
            subtype_by_id[stid] = (pname, stname)
            if stname:
                plant_strings.append(stname)
    plant_strings = list(dict.fromkeys(plant_strings))

    orders = list(
        db.orders.find(
            {"$or": [{"dealerOrder": False}, {"dealerOrder": {"$exists": False}}]},
            {
                "farmer": 1,
                "plantName": 1,
                "plantSubtype": 1,
                "orderId": 1,
                "orderStatus": 1,
                "numberOfPlants": 1,
                "additionalPlants": 1,
            },
        )
    )
    plant_name_by_id = {
        str(p["_id"]): (p.get("name") or "") for p in db.plantcms.find({}, {"name": 1})
    }

    orders_by_farmer: dict[str, list[dict]] = {}
    for o in orders:
        fid = o.get("farmer")
        if not fid:
            continue
        k = str(fid)
        orders_by_farmer.setdefault(k, []).append(o)

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_out = []

    for row in ws.iter_rows(min_row=1, values_only=True):
        cell = row[0] if row else None
        if cell is None or not str(cell).strip():
            continue
        line = str(cell).strip()
        if line.lower() == "particulars":
            continue

        am = amounts_from_row(row)
        tokens = parse_tokens(line)
        if len(tokens) < 3:
            rows_out.append(
                {
                    **am,
                    "excel_line": line,
                    "status": "SKIP_TOKENS",
                    "farmer_parse": "",
                    "village_parse": "",
                    "plant_parse": "",
                }
            )
            continue

        require_plant = False
        for ft, vt, pt in iter_splits(tokens):
            if not pt:
                continue
            pg = " ".join(pt)
            if best_plant_score(pg, plant_strings) >= MIN_PLANT:
                require_plant = True
                break

        best_row = None
        best_total = -1.0
        best_no_order = None
        best_no_order_total = -1.0

        for farmer_toks, village_toks, plant_toks in iter_splits(tokens):
            if require_plant and not plant_toks:
                continue
            farmer_guess = " ".join(farmer_toks)
            village_guess = " ".join(village_toks)
            plant_guess = " ".join(plant_toks)

            ps = best_plant_score(plant_guess, plant_strings) if plant_toks else 75.0
            if plant_toks and ps < MIN_PLANT:
                continue

            vg = norm(village_guess)
            hit = process.extractOne(
                vg,
                village_names_norm,
                scorer=fuzz.token_sort_ratio,
                score_cutoff=MIN_VILLAGE,
            )
            if not hit:
                continue
            best_v_score = float(hit[1])

            fg = norm(farmer_guess)
            farmer_name_list = [(fid, fd) for fid, fd in farmer_by_id.items()]
            f_names = [norm(fd["name"]) for _, fd in farmer_name_list]
            fhits = process.extract(
                fg,
                f_names,
                scorer=fuzz.token_sort_ratio,
                score_cutoff=MIN_FARMER,
                limit=80,
            )
            cand_farmer_ids = []
            for _, fs, idx in fhits:
                fid, fd = farmer_name_list[idx]
                cand_farmer_ids.append((fid, float(fs), fd))

            for fid, fs, fd in cand_farmer_ids:
                vn_sys = norm(fd["village_name"])
                vsys = max(
                    float(fuzz.ratio(vg, vn_sys)),
                    float(fuzz.partial_ratio(vg, vn_sys)),
                    float(fuzz.token_sort_ratio(vg, vn_sys)),
                )
                if vsys < MIN_VILLAGE:
                    continue

                olist = orders_by_farmer.get(fid, [])
                saw_order = False
                for ord_ in olist:
                    pid = str(ord_.get("plantName") or "")
                    sid = str(ord_.get("plantSubtype") or "")
                    p_sys = plant_name_by_id.get(pid, "")
                    st_pair = subtype_by_id.get(sid, ("", ""))
                    st_name = st_pair[1] if st_pair else ""
                    order_plant_blob = " ".join(x for x in (p_sys, st_name) if x)
                    if plant_toks:
                        os_score = best_plant_score(plant_guess, [p_sys, st_name, order_plant_blob])
                        if os_score < MIN_PLANT:
                            continue
                        total = fs * 0.4 + vsys * 0.35 + os_score * 0.25
                    else:
                        os_score = 0.0
                        total = fs * 0.55 + vsys * 0.45
                    saw_order = True
                    rec = {
                        **am,
                        "excel_line": line,
                        "status": "MATCHED",
                        "farmer_parse": farmer_guess,
                        "village_parse": village_guess,
                        "plant_parse": plant_guess or "(none)",
                        "score_farmer": round(fs, 1),
                        "score_village_parse_vs_db_village": round(best_v_score, 1),
                        "score_village_parse_vs_farmer_village": round(vsys, 1),
                        "score_plant": round(os_score, 1),
                        "matched_farmer_name": fd["name"],
                        "matched_village_db": fd["village_name"],
                        "orderId": ord_.get("orderId"),
                        "orderStatus": ord_.get("orderStatus"),
                        "order_plant": p_sys,
                        "order_subtype": st_name,
                        "numberOfPlants": ord_.get("numberOfPlants"),
                        "weighted_total": round(total, 2),
                    }
                    if total > best_total:
                        best_total = total
                        best_row = rec

                if (not olist or not saw_order) and plant_toks and ps >= MIN_PLANT:
                    fb_total = fs * 0.5 + vsys * 0.5
                    rec_no = {
                        **am,
                        "excel_line": line,
                        "status": "FARMER_MATCH_NO_ORDER",
                        "farmer_parse": farmer_guess,
                        "village_parse": village_guess,
                        "plant_parse": plant_guess,
                        "score_farmer": round(fs, 1),
                        "score_village_parse_vs_db_village": round(best_v_score, 1),
                        "score_village_parse_vs_farmer_village": round(vsys, 1),
                        "score_plant": round(ps, 1),
                        "matched_farmer_name": fd["name"],
                        "matched_village_db": fd["village_name"],
                        "orderId": "",
                        "orderStatus": "",
                        "order_plant": "",
                        "order_subtype": "",
                        "numberOfPlants": "",
                        "weighted_total": round(fb_total, 2),
                        "note": "no orders for this farmer id" if not olist else "no order passed plant threshold",
                    }
                    if fb_total > best_no_order_total:
                        best_no_order_total = fb_total
                        best_no_order = rec_no

        if best_row is None:
            if best_no_order is not None:
                rows_out.append(best_no_order)
            else:
                rows_out.append(
                    {
                        **am,
                        "excel_line": line,
                        "status": "NO_MATCH",
                        "farmer_parse": "",
                        "village_parse": "",
                        "plant_parse": "",
                    }
                )
        else:
            rows_out.append(best_row)

    wb.close()
    client.close()
    return rows_out


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else XLSX_DEFAULT
    try:
        rows_out = run_match(xlsx)
    except FileNotFoundError as e:
        print(e, file=sys.stderr)
        sys.exit(1)

    print("\t".join(OUTPUT_KEYS))
    for r in rows_out:
        print("\t".join(tsv_cell(r.get(k, "")) for k in OUTPUT_KEYS))

    matched = sum(1 for r in rows_out if r.get("status") == "MATCHED")
    print(f"\n# SUMMARY matched={matched} total={len(rows_out)}", file=sys.stderr)


if __name__ == "__main__":
    main()
