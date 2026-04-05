#!/usr/bin/env python3
"""
Build multi-sheet Excel: Final (first) | All | Matched | No_match (+ source sheets).
Uses prod DB via match_excel_particulars.run_match — same rules as CLI matcher.
"""
from __future__ import annotations

import sys
from pathlib import Path

SCRIPTS_DIR = Path(__file__).resolve().parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from match_excel_particulars import XLSX_DEFAULT, run_match

DEFAULT_OUT = SCRIPTS_DIR / "particulars_match_report_final.xlsx"

# Human-friendly column order; order_number = nursery orderId (numeric id in app)
SHEET_COLUMNS = [
    ("particulars", "excel_line"),
    ("status", "status"),
    ("order_number", None),  # derived from orderId
    ("debit_original", "excel_debit"),
    ("credit_original", "excel_credit"),
    ("matched_farmer_name", "matched_farmer_name"),
    ("matched_village_db", "matched_village_db"),
    ("order_plant", "order_plant"),
    ("order_subtype", "order_subtype"),
    ("number_of_plants", "numberOfPlants"),
    ("order_status", "orderStatus"),
    ("farmer_parse", "farmer_parse"),
    ("village_parse", "village_parse"),
    ("plant_parse", "plant_parse"),
    ("score_farmer", "score_farmer"),
    ("score_village_vs_catalog", "score_village_parse_vs_db_village"),
    ("score_village_vs_farmer", "score_village_parse_vs_farmer_village"),
    ("score_plant", "score_plant"),
    ("weighted_total", "weighted_total"),
    ("note", "note"),
]


def row_values(r: dict) -> list:
    oid = r.get("orderId")
    order_num = oid if oid not in (None, "") else ""
    out = []
    for header, key in SHEET_COLUMNS:
        if header == "order_number":
            out.append(order_num)
        else:
            v = r.get(key, "")
            if v is None:
                v = ""
            out.append(v)
    return out


def write_sheet(ws, rows: list[dict]) -> None:
    headers = [h for h, _ in SHEET_COLUMNS]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    for r in rows:
        ws.append(row_values(r))
    ws.freeze_panes = "A2"
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(str(h)), 12) + 2, 45)


def copy_source_sheet(wb: Workbook, source_xlsx: Path) -> None:
    from openpyxl import load_workbook

    src = load_workbook(source_xlsx, read_only=True, data_only=True)
    for name in src.sheetnames:
        n = name[:31] if len(name) > 31 else name
        while n in wb.sheetnames:
            n = (n[:28] + "...") if len(n) > 31 else n + "_"
        ws_new = wb.create_sheet(title=n)
        wss = src[name]
        for row in wss.iter_rows(values_only=True):
            ws_new.append(list(row))
    src.close()


def main() -> None:
    in_xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else XLSX_DEFAULT
    out_xlsx = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUT

    rows = run_match(in_xlsx)
    matched = [r for r in rows if r.get("status") == "MATCHED"]
    no_match = [r for r in rows if r.get("status") != "MATCHED"]

    wb = Workbook()
    ws_final = wb.active
    ws_final.title = "Final"
    write_sheet(ws_final, matched)

    ws_all = wb.create_sheet("All")
    write_sheet(ws_all, rows)

    ws_m = wb.create_sheet("Matched")
    write_sheet(ws_m, matched)

    ws_nm = wb.create_sheet("No_match")
    write_sheet(ws_nm, no_match)

    # Original workbook tabs (Sheet1, …) at the end for reference
    try:
        copy_source_sheet(wb, in_xlsx)
    except Exception as e:
        ws_note = wb.create_sheet("Source_error")
        ws_note.append(["Could not copy source sheets", str(e)])

    wb.save(out_xlsx)
    print(
        f"Wrote {out_xlsx} (Final+Matched={len(matched)}, all={len(rows)}, no_match={len(no_match)})"
    )


if __name__ == "__main__":
    main()
